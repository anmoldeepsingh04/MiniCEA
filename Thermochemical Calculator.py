# Project Name: Thermochemical Calculator
# Prepare by: Anmoldeep Singh, Senior(IVth) year, School of Mechanical Sciences, IIT Goa
# Date: 10/07/2021
# Python Version used: Python 3.8.2

# Note: Install openpyxl module before using this code:
# 1. Go to desktop and press Windows + X
# 2. Select Command prompt from the pop-up menu
# 3. Type: "pip install openpyxl" and press enter
# 4. After the installation is complete, type "exit" and press enter
# 5. Now you can run the program
# All the data for Heat of Formation are calculated at reference Temp of 298K

# Importing the relevant module
import openpyxl as op

# Importing the excel workbook for data
workbook = op.load_workbook(r'Data_File.xlsx')


# Defining a function for writing subscripts
def get_sub(x):
    normal = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+-=()"
    sub_s = "ₐ₈CDₑբGₕᵢⱼₖₗₘₙₒₚQᵣₛₜᵤᵥwₓᵧZₐ♭꜀ᑯₑբ₉ₕᵢⱼₖₗₘₙₒₚ૧ᵣₛₜᵤᵥwₓᵧ₂₀₁₂₃₄₅₆₇₈₉₊₋₌₍₎"
    res = x.maketrans(''.join(normal), ''.join(sub_s))
    return x.translate(res)


# Opening the relevant sheets from the workbook
CnHm = workbook['CnHm']
CO2 = workbook['CO2']
H2O = workbook['H2O']
N2 = workbook['N2']

print("Choose the fuel from the given list: ")  # Input for the type of fuel being used
for i in range(2, 27):
    print(i - 1, CnHm.cell(i, 1).value)
fuel_index = int(input("Enter the fuel index: "))
if fuel_index < 1 or fuel_index > 25:  # Checking for invalid inputs
    raise Exception("Invalid input, enter an index from given list.")  # Raising an error if invalid input is found

fuel = CnHm.cell(fuel_index + 1, 1).value  # Importing the fuel name
new_str = ''.join((ch if ch in '0123456789' else ' ') for ch in fuel)  # Getting the number of Carbon and Hydrogen atoms in the fuel
amount = [int(i) for i in new_str.split()]
if len(amount) == 2:
    amount.append(0)
print(amount)
no_C = amount[0]
no_H = amount[1]
no_O = amount[2]
T_ref = 298.0  # Defining the reference temperature at which the data is taken
if no_O == 0:
    compound_print = 'C{}H{}'.format(get_sub(str(no_C)), get_sub(str(no_H)))  # Fuel name for formatting
    compound = str('C{}H{}'.format(no_C, no_H))  # Fuel name
else:
    compound_print = 'C{}H{}O{}'.format(get_sub(str(no_C)), get_sub(str(no_H)), get_sub(str(no_O)))  # Fuel name for formatting
    compound = str('C{}H{}O{}'.format(no_C, no_H, no_O))  # Fuel name
print("The fuel chosen is:", compound_print)  # Printing the fuel chosen

row_count_CnHm = CnHm.max_row  # Number of rows in Fuel Sheet
h_f_comp = 0.0  # Heat of formation of fuel
T_adiabatic_actual = 0.0

# Getting data for the fuel
for i in range(2, row_count_CnHm + 1):
    comp_name = CnHm.cell(i, 1).value
    if compound == comp_name:
        h_f_comp = CnHm.cell(i, 3).value
        T_adiabatic_actual = CnHm.cell(i, 4).value

del_h_R = h_f_comp  # Net enthalpy on reactant side
row_count_prod = CO2.max_row  # Number of rows in Product Sheet

# Heat of formation of products
h_f_CO2 = 0.0
h_f_H2O = 0.0
h_f_N2 = 0.0

# Getting data for the products
for i in range(3, row_count_prod + 1):
    t = CO2.cell(i, 1).value
    if t == int(T_ref):
        h_f_CO2 = CO2.cell(i, 4).value
        h_f_H2O = H2O.cell(i, 4).value
        h_f_N2 = N2.cell(i, 4).value

# Estimated sensible heat of products at adiabatic temperature T(ad)
h_s_CO2 = 0.0
h_s_H2O = 0.0
h_s_N2 = 0.0

del_h_f_P = no_C * h_f_CO2 + (no_H / 2) * h_f_H2O + (no_C + (no_H / 4)) * 3.76 * h_f_N2  # Net heat of formation for products
net_h = del_h_R - del_h_f_P  # Net sensible heat for products at adiabatic temperature T(ad)
t_up = 0.0  # Upper limit of range for adiabatic temperature
t_down = 0.0  # Lower limit of range for adiabatic temperature
del_h_s_P_up = 0.0  # Sensible heat of products at upper limit temp
del_h_s_P_down = 0.0  # Sensible heat of products at lower limit temp

# Calculating the adiabatic temperature using given data
for i in range(3, row_count_prod + 1):
    t = CO2.cell(i, 1).value
    h_s_CO2 = CO2.cell(i, 3).value
    h_s_H2O = H2O.cell(i, 3).value
    h_s_N2 = N2.cell(i, 3).value
    del_h_s_P = no_C * h_s_CO2 + (no_H / 2) * h_s_H2O + (no_C + (no_H / 4)) * 3.76 * h_s_N2
    if del_h_s_P < net_h:
        del_h_s_P_down = del_h_s_P
        t_down = t
    if del_h_s_P > net_h:
        del_h_s_P_up = del_h_s_P
        t_up = t
        break

# Using linear interpolation between the upper and lower limit of adiabatic temperatures
T_adiabatic_calc = ((net_h - del_h_s_P_down) / (del_h_s_P_up - del_h_s_P_down)) * (t_up - t_down) + t_down

# Printing the combustion reaction equation
print("The combustion reaction is: ", end='')
if no_O != 0:
    print("{0} + {1}(O{2}+3.76N{3}) --> {4}CO{5} + {6}H{7}O + {8}N{9}".format(compound_print,
                                                                              no_C + no_H / 4.0 - no_O / 2.0, get_sub(str(2)),
                                                                              get_sub(str(2)), float(no_C), get_sub(str(2)),
                                                                              no_H / 2, get_sub(str(2)),
                                                                              round((no_C + no_H / 4 - no_O / 2) * 3.76, 3),
                                                                              get_sub(str(2))))
else:
    print("{0} + {1}(O{2}+3.76N{3}) --> {4}CO{5} + {6}H{7}O + {8}N{9}".format(compound_print,
                                                                              no_C + no_H / 4.0, get_sub(str(2)),
                                                                              get_sub(str(2)), float(no_C),
                                                                              get_sub(str(2)),
                                                                              no_H / 2, get_sub(str(2)),
                                                                              round((no_C + no_H / 4) * 3.76, 3),
                                                                              get_sub(str(2))))
# Printing the output
print("The lower limit of range for adiabatic temperature is: ", t_down, chr(176), "C", sep='')
print("The upper limit of range for adiabatic temperature is: ", t_up, chr(176), "C", sep='')
print("The actual adiabatic temperature is: ", T_adiabatic_actual, chr(176), "C", sep='')
print("The calculated adiabatic temperature is: ", round(T_adiabatic_calc, 3), chr(176), "C", sep='')
